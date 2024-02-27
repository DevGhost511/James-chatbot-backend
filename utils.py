from vectorizor import query_refiner
from scraper import get_chunks
from vectorizor import store_embeddings_in_pinecone, simple_generate, generate_surfing_image
from vectorizor import generate_answer
import pandas
import requests
import os
import openai
import simplejson as json
import openpyxl as xl 
from chunker import tiktoken_split
from dotenv import load_dotenv
load_dotenv()

# Set up STROM_GLASS_API_KEY
STROM_GLASS_API_KEY = '57387656-d281-11ee-8a07-0242ac130002-573876b0-d281-11ee-8a07-0242ac130002'


def generate_query(user_id, knowledge_name, chat_history, user_query):
    template = """
    The following is a friendly conversation between a human and an AI. The AI is talkative and provides lots of specific details from its context. If the AI does not know the answer to a question, it truthfully says it does not know current information but provides the latest information. If there is no conversation history, plz provide proper answer automatically based on the query.
    The response must be less than 300 words.\n
    Knowledge Base:{context}\n
    Chat history: {chat_history}\n
    Human: {human_input}\n
    Your response as Chatbot:"""
    
    print("Get query....")
    answer = generate_answer(query=user_query, user_id = user_id, knowledge_name=knowledge_name, latest_records=chat_history, template=template)

    return answer

def get_lat_long_for_beach(query):
    try:
        
        prompt = f"Provide the latitude and longitude and the name of the beach the user wants to go surfing in the query. This is the user's query:{query} Only respond in a format like this nothing else. Format: latitude, longitude, name of the place:34.0195, -118.4912, Arugam Bay"
        response = simple_generate(prompt)
        lat_long = response.split(',')
        latitude = lat_long[0].strip()
        longitude = lat_long[1].strip()
        beach = lat_long[2].strip()

        return latitude, longitude, beach
    except Exception as e:
        print(str(e))
        return None, None, None

def get_surf_conditions(lat, long):
    print("Strom glass api key ---->", STROM_GLASS_API_KEY)

    response = requests.get(
        'https://api.stormglass.io/v2/weather/point',
        params={
            'lat': lat,
            'lng': long,
            'params': ','.join(['waveHeight', 'waterTemperature', 'windSpeed', 'windDirection']),
        },
        headers={
            'Authorization': STROM_GLASS_API_KEY
        }
    )
    # Do something with response data.
    if response.status_code == 200:
        print(f"Received response from StormGlass: {response.json()}")
        return response.json()
    else:
        print(f"Failed to fetch data from StormGlass, Status Code={response.status_code}, Response={response.text}")
        return {'error': 'Failed to fetch data', 'status_code': response.status_code}

def generate_surf_response(query):
    try:
        surf_instructions, beach = get_weather_data_and_surf_instruction(query)
        print("OHeree-----")
        if surf_instructions and beach:
            hash_tag = generate_hashtags(surf_instructions)
            image_url = generate_surfing_image(surf_instructions, beach)
            print("Image URL =======>", image_url)
            return surf_instructions, image_url, hash_tag
    except Exception as e:
        print("Failed to generate the response:", e)
        return None, None, None

def get_weather_data_and_surf_instruction(query):
    latitude, longitude, beach = get_lat_long_for_beach(query)
    if latitude and longitude:
        surf_conditions = get_surf_conditions(latitude, longitude)
        print("Surfing =====>", surf_conditions)
        if 'error' not in surf_conditions:
            conditions_str = f"Wave Height: {surf_conditions['hours'][0]['waveHeight']['noaa']}m, Water Temperature: {surf_conditions['hours'][0]['waterTemperature']['noaa']}C, Wind Speed: {surf_conditions['hours'][0]['windSpeed']['noaa']}m/s, Wind Direction: {surf_conditions['hours'][0]['windDirection']['noaa']} degrees"
        else:
            conditions_str = "No conditions available"
    else:
        conditions_str = "No conditions available"

    surf_prompt = f"Write a facebook post update for the local surfers. You are a local surf expert at {beach}, considering the current wave conditions, select what size/type of board would be best for these conditions, and provide local tips for surfers based on the following conditions: {conditions_str}. Come up with other tips to say. Only say nice things about the local only if you mention them. The response must be less than 1500 characters"
    try:
        response = openai.chat.completions.create(
            model="gpt-4-turbo-preview",
            messages=[
                {"role": "system", "content": "You are a helpful assistant."},
                {"role": "user", "content": surf_prompt}
            ]
        )
        surf_instructions = response.choices[0].message.content
        return surf_instructions, beach

    except Exception as e:
        print("Failed to generate surfing content: %s", e)
        return None, None

def generate_hashtags(surf_instructions):
    hashtag_prompt = f"Based on the following content, generate relevant hashtags: {surf_instructions}"
    
    try:
        print("Generating hashtags")
        response = openai.chat.completions.create(
            model="gpt-4-turbo-preview",
            messages=[
                {"role": "system", "content": "You are a social media expert. Generate relevant hashtags based on the content."},
                {"role": "user", "content": hashtag_prompt}
            ]
        )
        hashtags = response.choices[0].message.content
        print("Generated Hashtags:\n%s", hashtags)
        return hashtags
    except Exception as e:
        print("Failed to generate hashtags: %s", e)
        return None

def generate_kb_from_xlsx(assistant_id, knowledge_id, path, api_key, index_name):
    excel = xl.load_workbook(path)
    sheetnames = excel.sheetnames
    res = {}
    for sheetname in sheetnames:
        data = pandas.read_excel("hospital cash price.xlsx", sheet_name=sheetname)
        json_data = data.to_json()
        res.update({sheetname: json_data})
    str_data = json.dumps(res)
    chunks = tiktoken_split(str_data)
    metadic ={'knowledge':knowledge_id, 'assistant':assistant_id}
    metalist = []
    ids = []
    
    for c in enumerate(chunks):
        metalist.append(metadic)
        ids.append(str(knowledge_id))

    store_embeddings_in_pinecone(chunks, metalist, api_key, ids, index_name)
    return len(chunks)

def generate_kb_from_file(assistant_id, knowledge_id, path, api_key, index_name):
    try:
        extension = path.split('.')[-1]
        extension = extension.lower()
        if extension == 'txt':
            count = generate_kb_from_txt(assistant_id=assistant_id, knowledge_id=knowledge_id, path=path, api_key=api_key, index_name=index_name)
            os.remove(path)
            return count
        if extension == 'xlsx' or extension == 'xls':
            count = generate_kb_from_xlsx(assistant_id, knowledge_id, path, api_key=api_key, index_name=index_name)
            os.remove(path)
            return count
        return -1
    except Exception as e:
        print(str(e))
        return -1

def generate_kb_from_txt(assistant_id, knowledge_id, path, api_key, index_name):
    try:
        with open(path, encoding='utf8') as f:
            content = f.read()
        chunks = tiktoken_split(content)
        metadic ={'knowledge':knowledge_id, 'assistant':assistant_id}
        metalist = []
        ids = []
        for index, chunk in enumerate(chunks):
            metalist.append(metadic)
            ids.append(str(knowledge_id)+"_"+str(index))

        store_embeddings_in_pinecone(chunks, metalist, api_key, ids, index_name)
        return len(chunks)
    except Exception as e:
        print(str(e))
        return -1

def generate_kb_from_url(assistant_id, knowledge_id, url, api_key, index_name):
    try:
        chunks = get_chunks(url)
        if chunks is None:
            return False
        metadic ={'knowledge':knowledge_id, 'assistant':assistant_id}
        print(chunks)
        metalist = []
        ids = []
        for index, chunk in enumerate(chunks):
            metalist.append(metadic)
            ids.append(str(knowledge_id)+"_"+str(index))
        store_embeddings_in_pinecone(chunks, metalist, api_key, ids, index_name)
        return len(chunks)
    except Exception as e:
        print(str(e))
        return -1

def get_response(query, prompt, latest_records, assistant_id):
    
    end = """
    Context:{context}
    Chat history:{chat_history}
    Human: {human_input}
    Assistant:"""
    template = prompt + end
    answer = generate_answer(query=query, assistant_id=assistant_id, latest_records=latest_records, template=template)
    
    return answer

def query_without_knowledge(chat_history, user_query):
    template = """The following is a friendly conversation between a human and an AI. The AI is talkative and provides lots of specific details from its context. If the AI does not know the answer to a question, it truthfully says it does not know.
        Context:{context}
        Chat history:{chat_history}
        Human: {human_input}
        AI:"""
    
    return ""
